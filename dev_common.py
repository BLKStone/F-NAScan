#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/11/30 0030 下午 21:20
# @Author  : BLKStone
# @Site    : http://blkstone.github.io
# @File    : dev_common.py
# @Software: PyCharm

# 外部依赖
from openpyxl import load_workbook
from openpyxl import Workbook
import validators

# 基本依赖
import socket
import datetime
import pickle


# 计时器类
# 用于对一段代码进行计时
class TimeKeeper(object):
    def __init__(self):
        self.start_time = None
        self.stop_time = None
        self.elapsed_time = None

    def start(self):
        self.start_time = datetime.datetime.now()

    def stop(self):
        self.stop_time = datetime.datetime.now()

    def elapsed(self):
        self.elapsed_time = self.stop_time - self.start_time
        message = '[*] elapsed ' + str(self.elapsed_time) + '...'
        print(message)
        return self.elapsed_time


# 对象Cooker
# 对 pickle 的进一步封装
# 便于将 python 对象存储为文件以及反向操作
class ObjectCooker(object):
    def __init__(self):
        pass

    def dump(self, python_object=None, path='default.pkl'):
        with open(path, 'wb') as f:
            pickle.dump(python_object, f)
            print('dumping... '+path)

    def load(self, path='default.pkl'):
        with open(path, 'rb') as f:
            python_object = pickle.load(f)
            print('loading... '+path)
        return python_object


# ListCooker
# 对读写文件进行封装
# 便于 txt 文件 与 python list 之间的相互转化
class ListCooker(object):
    def __init__(self):
        pass

    # 注意在读取时
    # 如果该行存在 '####' 则会被忽略
    def load_with_comment(self, path='default.txt'):
        python_list = []
        print(u'[info] comment loading... ' + path)
        with open(path, 'r') as f:
            for line in f.readlines():
                # '####' for comment
                if '####' in line:
                    continue
                if line[-1] == '\n':
                    python_list.append(line[:-1])
                else:
                    python_list.append(line)
        return python_list

    def load(self, path='default.txt'):
        python_list = []
        print(u'[info] loading... ' + path)
        with open(path, 'r') as f:
            for line in f.readlines():
                if line[-1] == '\n':
                    python_list.append(line[:-1])
                else:
                    python_list.append(line)
        return python_list

    def dump(self, python_list=None, path='default.txt'):
        with open(path, 'w') as f:
            print(u'[info] dumping... '+path)
            for ele in python_list:
                if ele[:-1] == '\n':
                    f.write(ele.strip().encode('utf-8'))
                else:
                    f.write(ele.strip().encode('utf-8')+'\n')
        f.close()


# FOFA 的 数据清洗辅助脚本
# 功能：
# 1. 从 fofa 数据中 对域名进行去重
# 2. 对 fofa蜘蛛 导出的数据中的 URL 进行协议标识补全
class FOFAWasher(object):
    def __init__(self):
        pass

    # 输入 一个url的list
    # 输出 域名list
    def distinct_domains(self, url_list):
        distinct = set()
        for url in url_list:
            try:
                distinct.add(url.split('/')[2])
            except Exception, e:
                print(e)
                continue

        result = list(distinct)
        return result

    # 用于清洗来自fofa爬虫的URL
    # 输入 url 的 list
    # 输出 url 的 list (前方带完整协议标记)
    def wash_url_list(self, urls):
        wash_urls = []
        for url in urls:
            if url[:5] == 'https':
                wash_urls.append(url)
            else:
                wash_urls.append('http://' + url)
        return wash_urls


# FNA-Scanner 辅助脚本
# 用于处理 FNA-Scanner 输出的结果
# 将端口扫描结果转化为可以作为Hydra输入的格式
class FNAParser(object):
    def __init__(self):
        pass

    def ftp_parse(self, input, output='ftp_hydra.txt'):
        '''
        Hydra 常见用法举例
        hydra -L dict/ftp_username.txt -P dict/top100.txt -e ns 192.168.0.1 ftp
        hydra -L dict/ftp_username.txt -P dict/top100.txt -e ns ftp://[10.15.44.172/24]/
        hydra -L dict/ftp_username.txt -P dict/top100.txt -e ns -M report/ftp.txt -o hydra_output.txt -Vv ftp

        '''

        lreader = ListCooker()
        fna_ftp_list = lreader.load(input)
        result = []
        for ele in fna_ftp_list:
            result.append(ele.split(' ')[0].strip())
        lreader.dump(result, output)


# 域名检查器
# 检查域名(字符串)是否合法
# 判断一个字符串是否是域名
class DomainChecker(object):
    def __init__(self):
        pass

    # 判断字符串是否为数字
    @staticmethod
    def is_number(string):
        try:
            num = int(string)
        except Exception,e:
            return False
        return True

    # 判断字符串是否为IP
    @staticmethod
    def is_ip_format(string):
        try:
            socket.inet_aton(string)
            # legal
        except Exception, e:
            # illegal
            return False
        return True

    # 判断一个字符串是否为域名
    @staticmethod
    def is_domain(string):
        # 直接利用 validators 来验证域名是否合法
        indicator = validators.domain(string)
        if indicator:  # indicator 为 True
            # print 'domain?', indicator, row[0].value
            return True
        else:
            return False

    # 是否是主机，或域名
    @staticmethod
    def temp_method(string):
        if string is None:
            # 排除 None 的情况
            # print "domain?", 'false'
            return False
        else:
            if not DomainChecker.is_number(string):
                # 不是数字 不是None
                if DomainChecker.is_ip_format(string):
                    # 该字符串是 ip
                    # print 'domain?', 'ip', row[0].value
                    return False
                else:
                    # 使用 validator 进一步确认域名
                    flag = validators.domain(string)
                    if flag:  # flag 为 True
                        # print 'domain?', flag, row[0].value
                        return True
                    else:
                        return False
            else:
                # 数字(端口号)排除
                # print 'domain?', 'number'
                return False


# DNS解析器
# 将域名转换为IP
class DNSAnalyzer(object):
    def __init__(self):
        pass

    @staticmethod
    def lookup(domain):
        ip = ''
        try:
            ip = socket.gethostbyname(domain)
        except Exception,e:
            return 'DNS_Resolve_Error'
        return ip


# 对资产发现平台导出的数据进行清洗的辅助类
class AssetsFinderWasher(object):
    def __init__(self):
        pass

    @staticmethod
    def webinfo_extrat(input_path, output_path):
        # 从 Excel (来自资产发现平台)中导出【web信息】区域的内容
        # 输出结果为单 sheet 的 Excel
        # 需要获取的核心数据为
        # URL(包含了域名信息) 机构名称 IP
        #
        # 举个例子
        # input_path = 'data/p1/assets2018.xlsx'
        # output_path = 'output.xlsx'

        # 存储需要写入输出excel的信息
        sheet_data = []

        # 将 Excel 读取到内存
        wb = load_workbook(filename=input_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print u'处理sheet...', sheet_name

            # 可以选择忽略某些特定名称的sheet
            # if sheet_name == 'sh.cn':
            #     print 'pass'
            #     continue

            webinfo_flag = False

            for row in ws.rows:
                # 存储行信息
                row_info = []

                # openpyxl 读取数据的常规操作
                # print 'row:', row
                # print 'cell:', row[0]
                # print 'value:', row[0].value

                if row[0].value is not None:

                    if u'web信息' in row[0].value:
                        # print row[0].value
                        webinfo_flag = True
                    else:
                        pass

                    if webinfo_flag:
                        # 读取行信息，封装在row_info(list)中
                        for cell in row:
                            # print cell.value,
                            row_info.append(cell.value)
                        # print ' '
                        print row_info

                        if len(row_info) != 3:
                            current_ip = '-'  # 默认值
                            if row_info[5] is None:
                                current_ip = row_info[0]
                            else:
                                try:
                                    if DomainChecker.is_ip_format(row_info[0]):
                                        # 忽略 URL 中包含 IP 的资产
                                        # 即
                                        # 类似 http://182.153.221.9:8080/index.php 的 URL 不会被统计
                                        # 类似 http://example.com:80/ 则会被纳入统计
                                        pass
                                    else:
                                        # 重组 URL
                                        # 写入 待输出的 sheet_data (list) 中
                                        if row_info[1] == '80':
                                            new_col = row_info[2] + '://' + row_info[0]
                                        else:
                                            new_col = row_info[2] + '://' + row_info[0] + ':' + row_info[1]
                                        sheet_data.append([row_info[0], new_col, row_info[6], current_ip])
                                except Exception, e:
                                    pass
                else:
                    pass

        # 将结果写进新文件
        new_wb = Workbook()
        new_ws = new_wb.active
        for ele in sheet_data:
            new_ws.append(ele)

        new_wb.save(output_path)









